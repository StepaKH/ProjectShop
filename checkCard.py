import openpyxl


def check_card_status(filename, phone_number):
    # Загружаем файл Excel
    wb = openpyxl.load_workbook(filename, read_only=False, keep_vba=True)
    ws = wb.active

    # Проходим по значениям в столбце "Моб. Телефон" и ищем совпадение с заданным номером телефона
    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        if str(row[9]) == phone_number:
            return 1  # Карта найдена
    return 0  # Карта не найдена


def create_card(filename, fio, phone, birthday):
    # Загружаем файл Excel
    wb = openpyxl.load_workbook(filename, read_only=False)
    ws = wb.active

    # Находим номер последней заполненной строки
    last_row = ws.max_row
    print(ws.cell(row=last_row, column=2).value)

    #Разделяем ФИО
    fio = fio.split()

    # Вписываем информацию в следующую свободную строку
    ws.cell(row=last_row + 1, column=1).value = ws.cell(row=last_row, column=1).value  # Вид ДК
    ws.cell(row=last_row + 1, column=2).value = int(ws.cell(row=last_row, column=2).value + 1)  # Штрих-код карты
    ws.cell(row=last_row + 1, column=2).number_format = '0'
    ws.cell(row=last_row + 1, column=3).value = fio[0]  # Фамилия
    ws.cell(row=last_row + 1, column=4).value = fio[1]  # Имя
    ws.cell(row=last_row + 1, column=5).value = fio[2]  # Отчество
    ws.cell(row=last_row + 1, column=6).value = birthday  # Дата рождения
    ws.cell(row=last_row + 1, column=10).value = phone  # Моб. Телефон

    # Сохраняем изменения в файле
    wb.save(filename)